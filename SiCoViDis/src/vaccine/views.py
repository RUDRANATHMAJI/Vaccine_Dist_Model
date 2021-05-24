from django.shortcuts import render

# Create your views here.

from django.shortcuts import render
from django.http import JsonResponse
import pandas as pd
from gurobipy import Model, GRB, quicksum
import gurobipy as gp
from itertools import product
import openpyxl
import requests
from openpyxl import load_workbook

def vaccine(request):
	return render(request,"vaccine.html",{})



def predict_chances(request):



    if request.POST.get('action') == 'post':

        # Receive data from client
        fac = float(request.POST.get('sepal_length'))
        
        #excel_file = (requests.POST.get('petal_length'))
        

        #excel_file = request.FILES["myfile"]

        #workbook = openpyxl.load_workbook(excel_file)
        
        #worksheet=workbook.active

       
        #excel_file = request.FILES["myfile"]

        #workbook = openpyxl.load_workbook(excel_file, data_only=True)
        #worksheet=workbook.active

        
        
        
        #Vaccine distribution model

        p=[i for i in range(26)]
        n=[i for i in range(26)]

        pot_fac = ['Hyderabad','Itanagar','Dispur', 'Patna', 'Raipur', 'Panaji', 'Gandhinagar', 'Chandigarh', 'Shimla', 'Ranchi','Bengaluru', 'Thiruvananthapuram', 'Bhopal','Mumbai', 'Imphal','Shillong', 'Aizawl', 'Kohima', 'Bhubaneswar', 'Jaipur', 'Gangtok', 'Chennai', 'Agartala', 'Lucknow', 'Dehradun', 'Kolkata'];
        customers = ['Hyderabad','Itanagar','Dispur', 'Patna', 'Raipur', 'Panaji', 'Ggandhinagar', 'Chandigarh', 'Shimla', 'Ranchi','Bengaluru', 'Thiruvananthapuram', 'Bhopal','Mumbai', 'Imphal','Shillong', 'Aizawl', 'Kohima', 'Bhubaneswar', 'Jaipur', 'Gangtok', 'Chennai', 'Agartala', 'Lucknow', 'Dehradun', 'Kolkata'];
        d=[39900000, 1736000, 35900000, 110500000, 30080000, 3700000, 71521000, 29808000, 7571000, 37000000, 68800000, 35122000, 86044000, 128400000, 3400000, 3652000, 1200000, 2200000, 47400000, 79500000, 672000, 77000000, 4100000, 231500000, 10400000, 99000000];
        dist = [[0, 2641,	2390, 1471, 779, 664, 1199, 1835,	1933, 1368, 569, 1251, 852, 709, 2866,	2479, 2850, 2732, 2732,	1554, 2027, 627, 2925, 1338,	1789, 1498 ],
        [2641, 0,	320,	1162, 1929, 3225, 2834,	2414, 2512, 1332, 3237,	3919, 2332, 2991, 467, 338, 675, 333, 1677,	2184, 771, 2937, 760, 1612, 2367, 1387],
        [2390, 320, 0,	938,	1705, 3000, 2609, 2190,	2287, 1107, 3013, 3695,	2107, 2766, 479, 91.3, 462, 344, 1452,	1959, 547, 2689, 537, 1388, 2143, 1163],
        [1471, 1162, 938, 0, 746, 2082, 1577, 1335, 1433, 323, 2094, 2776, 979, 1752, 1410, 1024, 1395, 1276, 766,	1106, 571, 2126, 1470, 533, 1288, 533],
        [779, 1929, 1705, 746, 0, 1361, 1157, 1493,	1591, 588, 1373, 2055, 629, 1215, 2180, 1794, 2165, 2046, 543, 1212, 1311, 1171, 2240, 787,	1446, 940],
        [664, 3225, 3000, 2082,	1361, 0, 1102,	2140, 2237, 1941, 594, 971, 1218,	571,	3439, 3053, 3423, 3305,	1760, 1593, 2600, 948, 3499,	1818, 2093, 2270],
        [1199, 2834, 2609,	1577, 1157, 1102, 0, 1147, 1245, 1614,	1512, 2202, 603, 553, 3089, 2702,	3073, 2955, 1694, 662, 2250,	1865, 3148, 1252, 1176, 1980],
        [1835, 2414, 2190,	1335, 1493, 2140, 1147,	0, 112, 1520, 2422, 3104, 1034, 1656, 2633,	2247, 2618, 2499, 1924,	510,	1794, 2454, 2693, 797, 223, 1778],
        [1933, 2512, 2287, 1433, 1591, 2237, 1245, 112,	0, 1618, 2519,	3201, 1131, 1754, 2731,	2344, 2715, 2597, 2022,	607,	1892, 2551, 2790, 894, 223, 1875],
        [1368, 1332, 1107,	323,	588,	1941, 1614, 1520, 1618,	0, 1962, 2331,	1016, 1706, 1577, 1191,	1562, 1443, 457, 1303, 738, 1619,	1638, 734, 1472, 408],
        [569, 3237, 3013, 2094,	1373, 594, 1512, 2422, 2519,	1962, 0, 683, 1443,	983,	3457, 3071, 3442, 3323,	1448, 2146, 2618, 347, 3517,	1930, 2380, 1883],
        [1251, 3919, 3695, 2776, 2055, 971, 2202, 3104,	3201, 2331, 683, 0,	2126, 1671, 4140, 3753,	4124, 4006, 1952, 2828,	3301, 724, 4199, 2612, 3062,	2393],
        [852, 2332, 2107, 979, 629, 1218,	603,	1034, 1131, 1016, 1443,	2126, 0, 777, 2500,	2113, 2484, 2210, 1172,	594, 1505, 1475, 2559, 661, 993, 1381],
        [709,	2991,	2766,	1752,	1215,	571,	553,	1656,	1754,	1706,	983,	1671,	777,	0,	3214,	2828,	3199,	3080,	1692,	1152,	2375,	1335,	3274,	1376,	1651,	2058],
        [2866,	467,	479,	1410,	2180,	3439,	3089,	2633,	2731,	1577,	3457,	4140,	2500,	3214,	0,	544,	401,	136,	1941,	2448,	1017,	3159,	538,	1876,	2631,	1543],
        [2479,	338,	91.3,	1024,	1794,	3053,	2702,	2247,	2344,	1191,	3071,	3753,	2113,	2828,	544,	0,	376,	411,	1537,	2044,	632,	2774,	451,	1472,	2021,	1139],
        [2850,	675,	462,	1395,	2165,	3423,	3073,	2618,	2715,	1562,	3442,	4124,	2484,	3199,	401,	376,	0,	537,	1908,	2415,	1002,	3145,	337,	1843,	2598,	1510],
        [2732,	333,	344,	1276,	2046,	3305, 2955,	2499,	2597,	1443,	3323,	4006, 2210,	3080,	136,	411,	537,	0,	1808,	2316,	884,	3027,	614,	1744,	2499,	1411],
        [2732,	1677,	1452,	766,	543,	1760,	1694,	1924,	2022,	457,	1448,	1952,	1172,	1692,	1941,	1537,	1908,	1808,	0,	1753,	1088,	1238,	1983,	1192,	1740,	442],
        [1554,	2184,	1959,	1106,	1212,	1593,	662,	510,	607,	1303,	2146,	2828,	594,	1152,	2448,	2044,	2415,	2316,	1753,	0,	1571,	2179,	2469,	572,	526,	1554],
        [2027,	771,	547,	571,	1311,	2600,	2250,	1794,	1892,	738,	2618,	3301,	1505,	2375,	1017,	632,	1002,	884,	1088,	1571,	0,	2322,	1077,	1020,	1568,	689],
        [627,	2937,	2689,	2126,	1171,	948,	1865,	2454,	2551,	1619,	347,	724,	1475,	1335,	3159,	2774,	3145,	3027,	1238,	2179,	2322,	0,	3550,	1963,	2413,	1674],
        [2925,	760,	537,	1470,	2240,	3499,	3148,	2693,	2790,	1638,	3517,	4199,	2559,	3274,	538,	451,	337,	614,	1983,	2469,	1077,	3550,	0,	1919,	2674,	1586],
        [1338,	1612,	1388,	533,	787,	1818,	1252,	797,	894,	734,	1930,	2612,	661,	1376,	1876,	1472,	1843,	1744,	1192,	572,	1020,	1963,	1919,	0,	550,	992],
        [1789,	2367,	2143,	1288,	1446,	2093,	1176,	223,	223,	1472,	2380,	3062,	993,	1651,	2631,	2021,	2598,	2499,	1740,	526,	1568,	2413,	2674,	550,	0,	1729],
        [1498,	1387,	1163,	533,	940,	2270,	1980,	1778,	1875,	408,	1883,	2393,	1381,	2058,	1543,	1139,	1510,	1411,	442,	1554,	689,	1674,	1586,	992,	1729,	0]];
        p1=fac;

        m=gp.Model('vaccdist')
        y=m.addVars(list(product(p,n)),vtype=GRB.BINARY)
        x=m.addVars(list(p),vtype=GRB.BINARY)

        m.addConstrs((gp.quicksum(y[i,j] for i in p)==1 for j in n));
        m.addConstr((gp.quicksum(x[i] for i in p) == p1));
        m.addConstrs((y[i,j] <= x[i] for j in n for i in p));

        m.setObjective(gp.quicksum(dist[i][j] * d[j] * y[i,j] for j in n for i in p ),GRB.MINIMIZE)

        m.optimize()

        v = m.getVars()
        p1=""
        p2=""
        p3=""
        p4=""
        p5=""
        p6=""
        p7=""
        p8=""
        p9=""
        p10=""
        p11=""
        p12=""
        p13=""
        p14=""
        p15=""
        p16=""
        p17=""
        p18=""
        p19=""
        p20=""
        p21=""
        p22=""
        p23=""
        p24=""
        p25=""
        p26=""
 

        if v[676].x == 1:
            p1 = ' Hyderabad'
        if v[677].x == 1:
            p2 = ' Itanagar'
        if v[678].x == 1:
            p3 = ' Dispur'
        if v[679].x ==1:
            p4 = ' Patna'
        if v[680].x == 1:
            p5 = ' Raipur'
        if v[681].x == 1:
            p6 = ' Panaji'
        if v[682].x ==1:
            p7 = ' Gandhinagar'
        if v[683].x == 1:
            p8 = ' Chandigarh'
        if v[684].x == 1:
            p9 = ' Shimla'
        if v[685].x ==1:
            p10 = ' Ranchi'
        if v[686].x == 1:
            p11 = ' Bengaluru'
        if v[687].x == 1:
            p12 = ' Thiruvananthapuram'
        if v[688].x ==1:
            p13 = ' Bhopal'
        if v[689].x ==1:
            p14 = ' Mumbai'
        if v[690].x == 1:
            p15 = ' Imphal'
        if v[691].x == 1:
            p16 = ' Shillong'
        if v[692].x ==1:
            p17 = ' Aizawl'
        if v[693].x ==1:
            p18 = ' Kohima'
        if v[694].x ==1:
            p19 = ' Bhubaneswar'
        if v[695].x == 1:
            p20 = ' Jaipur'
        if v[696].x == 1:
            p21 = ' Gangtok'
        if v[697].x ==1:
            p22 = ' Chennai'
        if v[698].x ==1:
            p23 = ' Agartala'
        if v[699].x == 1:
            p24 = ' Lucknow'
        if v[700].x == 1:
            p25 = ' Dehradun'
        if v[701].x ==1:
            p26 = ' Kolkata'

        ir=[384842,17446,235042,48816,327676,6490,737107,13088,69592,60500,197530,211011,46000,1070383,28051,14278,11275,9845,1556487,425766,55625,768118,45104,1056513,22662,157012]
        loc=[]
        i_rate = []
        for i in range(0,26):
            if float(v[i+676].x) == 1.0:
                loc.append(pot_fac[i])
                i_rate.append(int(ir[i]))
        i_rate2 = []
        for items in i_rate:
            i_rate2.append(items)

        i_rate.sort(reverse = True)
        
        res = {}
        for key in loc:
            for value in i_rate2:
                res[key] = value
                i_rate2.remove(value)
                break 
        sort_loc=[]



        sort_loc=list(dict(sorted(res.items(), key=lambda item: item[1])).keys())[::-1]

        for i in range(len(sort_loc)):
            sort_loc[i] = ' ' + sort_loc[i]  + ' '     



        return JsonResponse({'sepal_length':  sort_loc})                    

